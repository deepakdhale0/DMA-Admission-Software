from tkinter import *
from tkinter import ttk
from tkinter import scrolledtext
from tkinter import messagebox
import pymongo
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch

root = Tk()
root.geometry('1600x1000')
root.configure(bg='#f8f9fa')
root.title("Student Admission Form")

name_entry = StringVar()
student_con_entry = StringVar()
admission_date_entry = StringVar()
course_fee_entry = IntVar()
installment_entry_1 = IntVar()
installment_entry_2 = IntVar()
installment_entry_3 = IntVar()
installment_entry_4 = IntVar()
gender_combo = StringVar()
course_combo_1 = StringVar()
course_combo_2 = StringVar()
course_combo_3 = StringVar()
selected_year = StringVar()

first_year=['Element of Mining Technology', 'Element of Mining Geology' ]

second_year=['Basic Surveying','Mining Technology','Underground Coal Mining ','Mine Ventilation' ,'Rock Mechanics ','Engineering Mechanics' ,'Element of Mining Geology' ]

third_year=['Open Cast Mine ','Mine Management Legislation And Safety ','Method of Working Non Coal' ,'Mine Environment ','Advance Mine Surveying' ,'Mine Economics' ,'Mine Machinery']


frame1 = Frame(root, background='#003366', height=140, bd=15, relief=GROOVE)
frame1.pack(side=TOP, fill=BOTH)

Label(frame1, text='DMA',
      font=('Algerian', 25), anchor=CENTER, fg='white', bg='#003366').pack()
Label(frame1, text='Address : Jain Layout, Near Hanuman Mandir, Wani',
      font=('Arial', 15), anchor=CENTER, fg='white', bg='#003366').pack()
Label(frame1, text='Mobile: 9822543547 ',
      font=('Arial', 15), anchor=CENTER, fg='white', bg='#003366').pack()

frame2 = Frame(root, bg='#fdfdfd', bd=9, relief=GROOVE)
frame2.pack(side=TOP, fill=BOTH, expand=True)

frame3 = Frame(frame2, bg='#ffffff', height=900, width=400, bd=9, relief=GROOVE)
frame3.pack(side=LEFT, fill=BOTH, expand=True)

Label(frame3, text='STUDENT INFORMATION', font=('Arial', 16, 'bold'), bg='#ffffff', fg='#4E6688').grid(
    row=0, column=0, columnspan=2, pady=20)
Label(frame3, text='Name', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=1, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=name_entry, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=1, column=1, padx=20, pady=5)

Label(frame3, text='Student Contact', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=2, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=student_con_entry, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=2, column=1, padx=10, pady=5)

Label(frame3, text='Admission Date', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=3, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=admission_date_entry, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=3, column=1, padx=10, pady=5)

Label(frame3, text='Course Fee', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=6, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=course_fee_entry, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=6, column=1, padx=10, pady=5)

Label(frame3, text='1st Installment', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=7, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=installment_entry_1, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=7, column=1, padx=10, pady=5)

Label(frame3, text='2nd Installment', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=8, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=installment_entry_2, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=8, column=1, padx=10, pady=5)

Label(frame3, text='3rd Installment', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=9, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=installment_entry_3, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=9, column=1, padx=10, pady=5)

Label(frame3, text='4th Installment', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=10, column=0, padx=30, pady=5, sticky=W)
Entry(frame3, textvariable=installment_entry_4, font=('Arial', 11), bg='#f1f3f5', width=30).grid(row=10, column=1, padx=10, pady=5)

Label(frame3, text='Course Name', font=('Arial', 12), bg='#ffffff', fg='#343a40').grid(row=11, column=0, padx=30, pady=5, sticky=W)


frame_course = Frame(frame3, bg='#ffffff')
frame_course.grid(row=11, column=0, columnspan=2, padx=30, pady=10)

def display_subjects(subjects, year):
    subject_listbox.delete(0, END)
    for subject in subjects:
        subject_listbox.insert(END, subject)
    subject_title_label.config(text=f"{year} Subjects")
    selected_year.set(year)  

Button(frame_course, text='1st Year', command=lambda: display_subjects(first_year, "1st Year"),
       font=('Arial', 10), bg='#007bff', fg='white', width=10).grid(row=0, column=0, padx=5)

Button(frame_course, text='2nd Year', command=lambda: display_subjects(second_year, "2nd Year"),
       font=('Arial', 10), bg='#007bff', fg='white', width=10).grid(row=0, column=1, padx=5)

Button(frame_course, text='3rd Year', command=lambda: display_subjects(third_year, "3rd Year"),
       font=('Arial', 10), bg='#007bff', fg='white', width=10).grid(row=0, column=2, padx=5)

subject_title_label = Label(frame_course, text="", font=('Arial', 12, 'bold'), bg='#ffffff', fg='#343a40')
subject_title_label.grid(row=1, column=0, columnspan=3, pady=5)

subject_listbox = Listbox(frame_course, width=45, height=6, font=('Arial', 10), selectmode=MULTIPLE)
subject_listbox.grid(row=2, column=0, columnspan=3, pady=5)


frame4 = Frame(frame3, bg='#ffffff', bd=9, relief=GROOVE)
frame4.grid(row=12, column=0, columnspan=2, pady=20, padx=40)


frame5 = Frame(frame2, bg='white', height=600, width=800, bd=9, relief=GROOVE)
frame5.pack(side=RIGHT, fill=BOTH, expand=True)
def get_selected_subjects():
    selected_indices = subject_listbox.curselection()
    return [subject_listbox.get(i) for i in selected_indices]

frame6 = Frame(frame5, bd=9, relief=GROOVE, bg='white')
frame6.pack(side=TOP, fill=BOTH)
Label(frame6, text='Student Receipt', bd=9, font=('Arial', 14, 'bold'), bg='white', fg='black').pack(pady=10)

def bill():
    bill_text.delete('1.0', END)
    try:
        name = name_entry.get()
        student_con = student_con_entry.get()
        admission_date = admission_date_entry.get()
        course_fee = course_fee_entry.get()
        installment_1 = installment_entry_1.get()
        installment_2 = installment_entry_2.get()
        installment_3 = installment_entry_3.get()
        installment_4 = installment_entry_4.get()
        gender = gender_combo.get()
        selected_subjects = get_selected_subjects()
        year = selected_year.get()

        pending = course_fee - installment_1 - installment_2 - installment_3 - installment_4

        # Header
        bill_text.insert(END, "="*74 + "\n")
        bill_text.insert(END, f"{'DMA':^74}\n")
        bill_text.insert(END, f"{'Jain Layout, Ward No.6, Near Hanuman Mandir, Wani':^74}\n")
        bill_text.insert(END, f"{'Mobile: 9822543547':^74}\n")
        bill_text.insert(END, "="*74 + "\n")
        bill_text.insert(END, f"{'ADMISSION RECEIPT':^74}\n")
        bill_text.insert(END, "="*74 + "\n\n")

        # Student Info
        info_lines = [
            ("Name", name),
            ("Student Contact", student_con),
            ("Admission Date", admission_date),
            ("Gender", gender),
            ("Year", year)
        ]
        for label, value in info_lines:
            bill_text.insert(END, f"{label:<22}: {value}\n")

        # Subjects
        bill_text.insert(END, f"{'Subjects':<22}: ")
        if selected_subjects:
            bill_text.insert(END, f"{selected_subjects[0]}\n")
            for subject in selected_subjects[1:]:
                bill_text.insert(END, f"{'':<24}{subject}\n")
        else:
            bill_text.insert(END, "None\n")

        # Fee Details
        bill_text.insert(END, "-"*74 + "\n")
        fees = [
            ("Course Fee", course_fee),
            ("1st Installment", installment_1),
            ("2nd Installment", installment_2),
            ("3rd Installment", installment_3),
            ("4th Installment", installment_4),
            ("Pending Amount", pending)
        ]
        for label, amount in fees:
            bill_text.insert(END, f"{label:<22}: ₹{amount}\n")

        # Footer
        bill_text.insert(END, "-"*74 + "\n")
        bill_text.insert(END, f"\n{'Thank you for choosing DMA!':<74}\n")
        bill_text.insert(END, "="*74 + "\n")
        bill_text.insert(END, "\nSignature: " + "_"*40 + "\n")

    except Exception as e:
        bill_text.insert(END, f'\n[Error] Please enter all required information correctly.\nDetails: {e}')

bill_text = scrolledtext.ScrolledText(frame5, width=70, height=40, font=('Courier New', 10), bg='#f8f9fa')
bill_text.pack(padx=5, side=TOP)


# def database():
#     try:
#         con=pymongo.MongoClient("mongodb://localhost:27017/")
#         db=con['admin']
#         col=db['amplemind']
#         pending = course_fee_entry.get() - installment_entry_1.get() - installment_entry_2.get() - installment_entry_3.get() - installment_entry_4.get()

#         data={
#             'name':name_entry.get(),
#             'Student contact':student_con_entry.get(),
#             'Admission date':admission_date_entry.get(),
#             'Course Fee':course_fee_entry.get(),
#             '1st installment':installment_entry_1.get(),
#             '2st installment':installment_entry_2.get(),

#             'Gender':gender_combo.get(),
#             'Course':course_combo.get(),
#             'Pending Amount':pending
        
        
#             }
#         col.insert_one(data)
    
#         messagebox.showinfo('Student Info', 'Successfully submitted!')
#         print('data sucessfully enter in the database')
#     except Exception as e:
#         messagebox.showerror('Database Error', f'Failed to connect to database.\n{e}')
   

def save_bill_pdf():
    try:
        name = name_entry.get()
        student_con = student_con_entry.get()
        admission_date = admission_date_entry.get()
        course_fee = int(course_fee_entry.get())
        installment_1 = int(installment_entry_1.get())
        installment_2 = int(installment_entry_2.get())
        installment_3 = int(installment_entry_3.get())
        installment_4 = int(installment_entry_4.get())
        gender = gender_combo.get()
        selected_subjects = get_selected_subjects()
        year = selected_year.get()

        pending = course_fee - installment_1 - installment_2 - installment_3 - installment_4

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="Save Bill as PDF",
            initialfile=f"DMA_Receipt_{name.replace(' ', '_')}.pdf"
        )
        if not file_path:
            return

        c = canvas.Canvas(file_path, pagesize=A4)
        c.setFont("Courier", 10)
        width, height = A4
        y = height - 50

        def draw_line(text):
            nonlocal y
            c.drawString(40, y, text)
            y -= 15

        draw_line("=" * 85)
        draw_line(f"{'DMA':^85}")
        draw_line(f"{'Jain Layout, Ward No.6, Near Hanuman Mandir, Wani':^85}")
        draw_line(f"{'Mobile: 9822543547':^85}")
        draw_line("=" * 85)
        draw_line(f"{'ADMISSION RECEIPT':^85}")
        draw_line("=" * 85)
        draw_line("")

        draw_line(f"{'Name':<22}: {name}")
        draw_line(f"{'Student Contact':<22}: {student_con}")
        draw_line(f"{'Admission Date':<22}: {admission_date}")
        draw_line(f"{'Gender':<22}: {gender}")
        draw_line(f"{'Year':<22}: {year}")

        draw_line(f"{'Subjects':<22}: " + (selected_subjects[0] if selected_subjects else "None"))
        for subject in selected_subjects[1:] if selected_subjects else []:
            draw_line(f"{'':<24}{subject}")

        draw_line("-" * 85)
        draw_line(f"{'Course Fee':<22}: ₹{course_fee}")
        draw_line(f"{'1st Installment':<22}: ₹{installment_1}")
        draw_line(f"{'2nd Installment':<22}: ₹{installment_2}")
        draw_line(f"{'3rd Installment':<22}: ₹{installment_3}")
        draw_line(f"{'4th Installment':<22}: ₹{installment_4}")
        draw_line(f"{'Pending Amount':<22}: ₹{pending}")
        draw_line("-" * 85)
        draw_line("")
        draw_line("Thank you for choosing DMA!")
        draw_line("=" * 85)
        draw_line("")
        draw_line("Signature: " + "_" * 40)

        c.save()
        messagebox.showinfo("Success", f"PDF saved successfully:\n")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to save PDF.\n{e}")

def export_to_excel():
    file_path = "student class data.xlsx"

    name = name_entry.get()
    student_con = student_con_entry.get()
    admission_date = admission_date_entry.get()
    course_fee = course_fee_entry.get()
    installment_1 = installment_entry_1.get()
    installment_2 = installment_entry_2.get()
    installment_3 = installment_entry_3.get()
    installment_4 = installment_entry_4.get()

    selected_subjects = get_selected_subjects()
    subject_str = ", ".join(selected_subjects)

    pending = course_fee - installment_1 - installment_2 - installment_3 - installment_4

    headers = [
        "Name", "Student Contact",  "Admission Date",
        "Subjects",
        "Course Fee", "1st Installment", "2nd Installment", "3rd Installment", "4th Installment", "Pending Amount"
    ]

    values = [
        name, student_con,  admission_date, subject_str,
        course_fee, installment_1, installment_2, installment_3, installment_4, pending
    ]

    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Admission Details"
            ws.append(headers)  # Add headers only once

        ws.append(values)
        wb.save(file_path)
        messagebox.showinfo("Success", "Data added to Excel successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export data to Excel.\n{e}")


def clear_form():
    name_entry.set("")
    student_con_entry.set("")
    admission_date_entry.set("")
    course_fee_entry.set(0)
    installment_entry_1.set(0)
    installment_entry_2.set(0)
    installment_entry_3.set(0)
    installment_entry_4.set(0)
    course_combo_1.set("Select Course")
    course_combo_2.set("Select Course")
    course_combo_3.set("Select Course")
    
    bill_text.delete('1.0', END)
    
    subject_listbox.selection_clear(0, END)
    subject_listbox.delete(0, END)
    subject_title_label.config(text="")  

                                                                    
Button(frame4, text='PDF',command=save_bill_pdf, font=('Arial', 12), bg='#004080', fg='white', activebackground='#003366', width=15).pack(side=LEFT, padx=10,pady=7)
Button(frame4, text='Print Bill', command=bill, font=('Arial', 12), bg='#004080', fg='white', activebackground='#003366', width=15).pack(side=LEFT, padx=10,pady=7)
Button(frame4, text='clear',command=clear_form, font=('Arial', 12), bg='#004080', fg='white', activebackground='#003366', width=18).pack(side=LEFT, padx=10,pady=7)
Button(frame4, text='Add to Excel', command=export_to_excel, font=('Arial', 12), bg='#004080', fg='white', activebackground='#003366', width=15).pack(side=LEFT, padx=10, pady=7)


root.mainloop()



      